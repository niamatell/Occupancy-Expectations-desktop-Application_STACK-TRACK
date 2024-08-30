import sys
from src.ui_interface import *
import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import hashlib
import secrets
import datetime
from datetime import date 
import re
import matplotlib as mpl
from datetime import datetime, timedelta
import uuid
from fpdf import FPDF 
from PySide6.QtWidgets import QTableWidgetItem
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from Custom_Widgets import *
from collections import defaultdict
from Custom_Widgets.QAppSettings import QAppSettings
from PySide6.QtWidgets import QFileDialog
from PySide6.QtWidgets import QVBoxLayout


class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)


        self.ui.mainPages.currentChanged.connect(self.clear_error_messages)
        self.ui.loginBtn.clicked.connect(self.login)
        self.ui.goToSigninBtn.clicked.connect(self.go_to_signup_page)
        self.ui.signupBtn.clicked.connect(self.signup)
        self.ui.goToLoginBtn.clicked.connect(self.go_to_login_page)
        self.ui.userBtn.clicked.connect(self.logout)
        self.ui.eyeBtn.clicked.connect(lambda: self.toggle_password_visibility(self.ui.password, self.ui.eyeBtn))
        self.ui.eyeBtn_2.clicked.connect(lambda: self.toggle_password_visibility(self.ui.newPassword, self.ui.eyeBtn_2))
        self.ui.eyeBtn_3.clicked.connect(lambda: self.toggle_password_visibility(self.ui.currentPassword, self.ui.eyeBtn_3))
        self.ui.eyeBtn_4.clicked.connect(lambda: self.toggle_password_visibility(self.ui.passEmail, self.ui.eyeBtn_4))
        self.ui.eyeBtn_5.clicked.connect(lambda: self.toggle_password_visibility(self.ui.signPassword, self.ui.eyeBtn_5))
        self.ui.changePassBtn.clicked.connect(self.change_password)
        self.ui.changeEmailBtn.clicked.connect(self.change_email)
        self.ui.stackedWidget.currentChanged.connect(self.on_page_changed)

        self.ui.mainPages.currentChanged.connect(self.on_page_changed_3)
        self.ui.tableWidget_2.itemChanged.connect(self.handle_item_changed)
        self.ui.exportBtn.clicked.connect(self.export_table_data)

        loadJsonStyle(self, self.ui, jsonFiles = {
            "json-styles/style.json"
            }) 

        self.ui.stackedWidget_2.setCurrentIndex(0)
        self.ui.stackedWidget.setCurrentIndex(0)
        self.ui.mainPages.setCurrentIndex(0)

        self.user_initiated_change = False
        self.show() 
        QAppSettings.updateAppSettings(self)

    def toggle_password_visibility(self, password_edit, eye_button):
        if password_edit.echoMode() == QLineEdit.Normal:
            password_edit.setEchoMode(QLineEdit.Password)
            eye_button.setIcon(QIcon("Qss/icons/Icons/feather/eye.png"))
        else:
            password_edit.setEchoMode(QLineEdit.Normal)
            eye_button.setIcon(QIcon("Qss/icons/Icons/feather/eye-off.png"))
    #Manage changing pages
    def on_page_changed(self, index):
        if index == 1:
            self.fetch_data_and_calculate_expected_occupancy()
            self.generate_pie_chart()
            self.generate_occupancy_variation_graph()
            self.generate_yard_density_chart()

    def on_page_changed_3(self, index):
        if index == 0:
            self.update_charts()
   
    def clear_error_messages(self):
        self.ui.errorMessage.setText("")
        self.ui.errorMessage_2.setText("")
        self.ui.errorMessage_3.setText("")
        self.ui.errorMessage_4.setText("")
        self.ui.successMessage.setText("")
        self.ui.successMessage_2.setText("")

    def go_to_signup_page(self):
        self.ui.stackedWidget_2.setCurrentIndex(1)
        self.ui.errorMessage_2.setText("")

    def go_to_login_page(self):
        self.ui.stackedWidget_2.setCurrentIndex(0)
        self.ui.errorMessage.setText("")
    
        
    #Create connection with db
    def create_connection(self):
        conn = pyodbc.connect(
            r'DRIVER={SQL Server};SERVER={DESKTOP-GEKSV3P\SQLEXPRESS01};DATABASE={egt_db};Trusted_Connection=yes;'
        )
        return conn

    #Manage Login page
    def login(self):
        username = self.ui.username.text()
        password = self.ui.password.text()
        if self.authenticate_user(username, password):
            self.ui.stackedWidget.setCurrentIndex(1)
            self.ui.errorMessage.setText("")
        else:
            print("Login failed. Invalid username or password.")
            self.ui.errorMessage.setText("Username or password incorrect.")   

    def signup(self):
        try:
            full_name = self.ui.fullName.text().strip()
            username = self.ui.signUsername.text().strip()
            email = self.ui.email.text().strip()
            password = self.ui.signPassword.text().strip()

            # Check for empty fields
            if not full_name or not username or not email or not password:
                self.ui.errorMessage_2.setText("All fields are required")
                return False
            
            # Check if the email is valid
            if not self.is_valid_email(email):
                self.ui.errorMessage_2.setText("Email is not valid")
                return False
            
            # Check if the password length is at least 8 characters
            if len(password) < 8:
                self.ui.errorMessage_2.setText("Password should be 8 characters or more")
                return False
            
            # Check if the username or email already exists in the database
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT username, email FROM users WHERE username = ? OR email = ?", (username, email))
            existing_users = cursor.fetchall()

            for user in existing_users:
                if user[0] == username:
                    self.ui.errorMessage_2.setText("Username already exists")
                    cursor.close()
                    conn.close()
                    return False
                if user[1] == email:
                    self.ui.errorMessage_2.setText("Email already exists")
                    cursor.close()
                    conn.close()
                    return False

            # Hash the password
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            
            # Generate session token
            session_token = self.generate_session_token()
            
            # Insert the new user into the database
            cursor.execute(
                "INSERT INTO users (full_name, username, email, password, session_token) VALUES (?, ?, ?, ?, ?)",
                (full_name, username, email, hashed_password, session_token)
            )
            conn.commit()
            cursor.close()
            conn.close()
            
            # Clear error message and return success
            self.ui.errorMessage.setText("")
            print("User registered successfully.")
            self.ui.stackedWidget_2.setCurrentIndex(0)
            return True
        except Exception as e:
            self.ui.errorMessage.setText("An error occurred during registration")
            return False
        
    def is_valid_email(self, email):
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(email_regex, email) is not None
    
    def authenticate_user(self, username, password):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT password FROM users WHERE username = ?", (username,))
            result = cursor.fetchone()
            if result:
                hashed_password_from_db = result[0]
                hashed_password_entered = hashlib.sha256(password.encode()).hexdigest()
                if hashed_password_from_db == hashed_password_entered:
                    session_token = str(uuid.uuid4())
                    session_expiry = datetime.now() + timedelta(days=1)
                    cursor.execute("UPDATE users SET session_token = ?, session_expiry = ? WHERE username = ?", (session_token, session_expiry, username))
                    conn.commit()
                    return True  
            cursor.close()
            conn.close()
            return False  
        except Exception as e:
            print(f"Error authenticating user: {e}")
            return False
   
    def generate_session_token(self):
        session_token = secrets.token_hex(16)  
        return session_token
         
    def update_session(self, username, session_token, session_expiry):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE users SET session_token = ?, session_expiry = ? WHERE username = ?", (session_token, session_expiry, username))
            conn.commit()

            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error updating session: {e}")

    def logout(self):
        self.ui.username.clear()
        self.ui.password.clear()
        self.ui.stackedWidget.setCurrentIndex(0)
        self.clear_error_messages()

    def change_password(self):
        current_password = self.ui.currentPassword.text()
        new_password = self.ui.newPassword.text()
        username = self.ui.username.text()  

        if len(new_password) < 8:
            self.ui.successMessage.setText("")
            self.ui.errorMessage_3.setText("Password should be 8 characters or more")
            return False

        if not self.verify_current_password(username, current_password):
            self.ui.successMessage.setText("")
            self.ui.errorMessage_3.setText("Current password incorrect")
            return False

        hashed_new_password = hashlib.sha256(new_password.encode()).hexdigest()
        if self.update_password(username, hashed_new_password):
            self.update_password(username, hashed_new_password)
            self.ui.currentPassword.clear()
            self.ui.newPassword.clear()
            self.ui.errorMessage_3.setText("")
            print("Password updated successfully.")
            self.ui.successMessage.setText("Password changed successfully.")
            return True
        else:
            self.ui.errorMessage_3.setText("An error occurred while updating the password")
            return False
        
    def change_email(self):
        current_password = self.ui.passEmail.text()
        new_email = self.ui.newEmail.text()
        username = self.ui.username.text() 

        if not self.is_valid_email(new_email):
            self.ui.successMessage_2.setText("")
            self.ui.errorMessage_4.setText("Email is not valid")
            return False
        if not self.verify_current_password(username, current_password):
            self.ui.successMessage_2.setText("")
            self.ui.errorMessage_4.setText("Current password incorrect")
            return False
        if self.update_email(username, new_email):
            self.ui.passEmail.clear()
            self.ui.newEmail.clear()
            self.ui.errorMessage_4.setText("")  
            print("Email updated successfully.")
            self.ui.successMessage_2.setText("Email changed successfully.")
            return True
        else:
            self.ui.errorMessage_4.setText("An error occurred while updating the email")
            return False

    def verify_current_password(self, username, current_password):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT password FROM users WHERE username = ?", (username,))
            result = cursor.fetchone()
            if result:
                hashed_password_from_db = result[0]
                hashed_current_password = hashlib.sha256(current_password.encode()).hexdigest()
                return hashed_password_from_db == hashed_current_password
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error verifying current password: {e}")
            return False

    def update_password(self, username, hashed_new_password):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE users SET password = ? WHERE username = ?", (hashed_new_password, username))
            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            print(f"Error updating password: {e}")
            return False

    def update_email(self, username, new_email):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE users SET email = ? WHERE username = ?", (new_email, username))
            conn.commit()
            cursor.close()
            conn.close()
            return True
        except Exception as e:
            print(f"Error updating email: {e}")
            return False


    #Exporting files :
    def export_table_data(self):
        rowCount = self.ui.tableWidget.rowCount()
        columnCount = self.ui.tableWidget.columnCount()
        data = []
        headers = []
        for column in range(columnCount):
            headers.append(self.ui.tableWidget.horizontalHeaderItem(column).text())
        for row in range(rowCount):
            rowData = []
            for column in range(columnCount):
                item = self.ui.tableWidget.item(row, column)
                if item is not None:
                    rowData.append(item.text())
                else:
                    rowData.append('')
            data.append(rowData)
        df = pd.DataFrame(data, columns=headers)
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx);;PDF Files (*.pdf)", options=options)
        
        if fileName:
            if fileName.endswith('.xlsx'):
                df.to_excel(fileName, index=False)
                workbook = load_workbook(fileName)
                worksheet = workbook.active
                header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") 
                for col_num, col in enumerate(worksheet.iter_cols(1, worksheet.max_column), 1):
                    max_length = len(headers[col_num - 1])
                    for cell in col:
                        if cell.row == 1:  
                            cell.fill = header_fill
                            cell.font = Font(bold=True)
                        else:
                            max_length = max(max_length, len(str(cell.value) if cell.value is not None else ""))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[col[0].column_letter].width = adjusted_width
            
                workbook.save(fileName)
            elif fileName.endswith('.pdf'):
                self.export_to_pdf(fileName)

    def export_to_pdf(self, file_path):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=8)
        title = f"Forecast Data of {date.today()}"
        pdf.set_font("Arial", 'B', size=8)
        pdf.cell(0, 10, title, 0, 1, 'C')
        pdf.ln(10)
        pdf.set_font("Arial", size=6)
        
        headers = [self.ui.tableWidget.horizontalHeaderItem(col).text() for col in range(self.ui.tableWidget.columnCount())]
        col_widths = [pdf.get_string_width(header) + 10 for header in headers]

        for row in range(self.ui.tableWidget.rowCount()):
            for col in range(self.ui.tableWidget.columnCount()):
                item = self.ui.tableWidget.item(row, col)
                cell_text = item.text() if item is not None else ""
                col_widths[col] = max(col_widths[col], pdf.get_string_width(cell_text) + 10)

        for col, header in enumerate(headers):
            pdf.set_fill_color(173, 216, 230)  # Light blue color
            pdf.cell(col_widths[col], 10, header, 1, 0, 'C', fill=True)
        pdf.ln()
        
        for row in range(self.ui.tableWidget.rowCount()):
            for col in range(self.ui.tableWidget.columnCount()):
                item = self.ui.tableWidget.item(row, col)
                cell_text = item.text() if item is not None else ""
                pdf.cell(col_widths[col], 10, cell_text, 1)
            pdf.ln()
        pdf.output(file_path)
        print(f'Table exported to {file_path} successfully!')

    #Manage charts on homePage
    def update_charts(self):
        self.generate_pie_chart()
        self.generate_occupancy_variation_graph()
        self.generate_yard_density_chart()

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()

    def generate_pie_chart(self):
        try:
            actual_occupancy = float(self.ui.tableWidget_2.item(0, 0).text())
            operational_capacity = float(self.ui.tableWidget_2.item(0, 1).text())
            empty_space = operational_capacity - actual_occupancy
            labels = ['Occupied Space', 'Empty Space']
            sizes = [actual_occupancy, empty_space]
            colors = ['#EFBA2E', 'lightskyblue']
            fig = Figure(figsize=(5, 3))
            ax = fig.add_subplot(111)
            ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=140, textprops={'fontsize': 8})
            ax.axis('equal') 
            canvas = FigureCanvas(fig)
            layout = self.ui.pieChartWidget.layout()
            if layout is None:
                layout = QVBoxLayout()
                self.ui.pieChartWidget.setLayout(layout)
            self.clear_layout(layout) 
            layout.addWidget(canvas)
        except Exception as e:
            print(f"Error generating pie chart: {e}")

    def generate_occupancy_variation_graph(self):
        try:
            occupancy_by_day = defaultdict(float)
            for row in range(self.ui.tableWidget.rowCount()):
                etadate = self.ui.tableWidget.item(row, 0).text()
                expected_occupancy = float(self.ui.tableWidget.item(row, 6).text())
                day = etadate.split()[0]  
                occupancy_by_day[day] = expected_occupancy  

            day_labels = list(occupancy_by_day.keys())
            occupancies = list(occupancy_by_day.values())

            mpl.rcParams.update(mpl.rcParamsDefault)
            plt.style.use("ggplot")

            fig = Figure(figsize=(10, 4))
            ax = fig.add_subplot(111)
            ax.plot(day_labels, occupancies, linestyle='-', color='lightskyblue')
            ax.grid(True)
            ax.tick_params(axis='x', rotation=30, which='major', width=1, length=5, pad=10, labelsize=8) 
            ax.tick_params(axis='y', rotation=0, which='major', width=1, length=5, pad=10, labelsize=8) 
            fig.tight_layout()

            canvas = FigureCanvas(fig)

            layout = self.ui.lineGraphWidget.layout()
            if layout is None:
                layout = QVBoxLayout(self.ui.lineGraphWidget)
                self.ui.lineGraphWidget.setLayout(layout)
            self.clear_layout(layout)  # Clear the previous chart
            layout.addWidget(canvas)
        except Exception as e:
            print(f"Error generating occupancy variation graph: {e}")

    def generate_yard_density_chart(self):
        try:
            yard_density = float(self.ui.tableWidget_2.item(0, 2).text())
            if not hasattr(self.ui, 'progressBarYardDensity'):
                self.ui.progressBarYardDensity = QProgressBar(self.ui.densityChartWidget)
                layout = self.ui.densityChartWidget.layout()
                if layout is None:
                    layout = QVBoxLayout(self.ui.densityChartWidget)
                    self.ui.densityChartWidget.setLayout(layout)
                layout.addWidget(self.ui.progressBarYardDensity)

            self.ui.progressBarYardDensity.setRange(0, 100)
            self.ui.progressBarYardDensity.setValue(yard_density)
            self.ui.progressBarYardDensity.setFormat(f'{yard_density:.1f}%')

            border_style = """
                QProgressBar {
                    border: 2px solid #4A4A4A;
                    border-radius: 5px;
                    text-align: center;
                    height: 30px;
                }
            """
            
            if yard_density < 40:
                color = "green"
            elif yard_density < 70:
                color = "#EFBA2E"
            else:
                color = "red"
            
            chunk_style = f"""
                QProgressBar::chunk {{
                    background-color: {color};
                    width: 10px;
                    margin: 1px;
                    border-radius: 3px;
                }}
            """
            
            self.ui.progressBarYardDensity.setStyleSheet(border_style + chunk_style)
        except Exception as e:
            print(f"Error updating yard density progress bar: {e}")

    #Manage reportsPage
    def update_db(self, actual_occupancy, operational_capacity):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            yard_density = actual_occupancy/operational_capacity*100
            cursor.execute("""
                INSERT INTO state (actual_occupancy, operational_capacity, yard_density, change_date)
                VALUES (?, ?, ?, ?)
            """, (actual_occupancy, operational_capacity, yard_density, current_date))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error updating DB: {e}")
    
    def set_state_table(self, actual_occupancy, operational_capacity):
        self.ui.tableWidget_2.clearContents()
        self.ui.tableWidget_2.setRowCount(1)  
        yard_density = (actual_occupancy / operational_capacity) * 100
        self.ui.tableWidget_2.setItem(0, 0, QTableWidgetItem(str(round(actual_occupancy,0))))
        self.ui.tableWidget_2.setItem(0, 1, QTableWidgetItem(str(operational_capacity)))
        self.ui.tableWidget_2.setItem(0, 2, QTableWidgetItem(str(round(yard_density,2))))
    
    def insert_or_update_calculated_data(self, etadate, vessel_service, vessel_name, load_quantity, discharge_quantity, expected_teu, expected_occupancy, PrimaryKey):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT COUNT(*) FROM Data_forecast 
                WHERE etadate = ? AND PrimaryKey = ?
            """, (etadate, PrimaryKey))
            exists = cursor.fetchone()[0]

            if exists:
                cursor.execute("""
                    UPDATE Data_forecast 
                    SET load_quantity = ?, discharge_quantity = ?, expected_teu = ?, expected_occupancy = ?
                    WHERE etadate = ? AND vessel_service = ? AND vessel_name = ? AND PrimaryKey = ?
                """, (load_quantity, discharge_quantity, expected_teu, expected_occupancy, etadate, vessel_service, vessel_name, PrimaryKey))
            else:
                cursor.execute("""
                    INSERT INTO Data_forecast (etadate, vessel_service, vessel_name, load_quantity, discharge_quantity, expected_teu, expected_occupancy, PrimaryKey)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (etadate, vessel_service, vessel_name, load_quantity, discharge_quantity, expected_teu, expected_occupancy, PrimaryKey))

            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error inserting or updating calculated data: {e}")

    def fetch_data_and_calculate_expected_occupancy(self):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT TOP 1 actual_occupancy, operational_capacity FROM state ORDER BY change_date DESC")
            actual_occupancy, operational_capacity = cursor.fetchone()
            
            cursor.execute("""
                        SELECT ETADATE, VesselService, VesselName, LoadQuantity, DischargeQuantity, PrimaryKey 
                        FROM Vessels_data 
                        WHERE VoyageStatus = 'P'""")
            data = cursor.fetchall()
            updated_values = {}
            self.ui.tableWidget.clearContents()
            self.ui.tableWidget.setRowCount(len(data))
            self.set_state_table(actual_occupancy, operational_capacity)
            
            for row_index, row_data in enumerate(data):
                etadate, vessel_service, vessel_name, load_quantity, discharge_quantity, PrimaryKey = row_data
                loadQuantity = float(load_quantity)
                dischargeQuantity = float(discharge_quantity)
                

                if loadQuantity == 0.0 or dischargeQuantity == 0.0:
                    if vessel_service in updated_values:
                        loadQuantity, dischargeQuantity = updated_values[vessel_service]
                        print(f"Using updated values for {vessel_service}: {loadQuantity} // {dischargeQuantity}")
                    else:
                        cursor.execute("""
                            SELECT TOP 3 LoadQuantity, DischargeQuantity 
                            FROM Vessels_data 
                            WHERE VesselService = ? AND ETADATE < ? 
                            ORDER BY ETADATE DESC
                            """, (vessel_service, etadate))
                        last_three_visits = cursor.fetchall()

                        if last_three_visits:
                            last_three_visits = [(float(visit[0]), float(visit[1])) for visit in last_three_visits]
                            average_load_quantity = sum(visit[0] for visit in last_three_visits) / len(last_three_visits)
                            average_discharge_quantity = sum(visit[1] for visit in last_three_visits) / len(last_three_visits)
                        else:
                            average_load_quantity = 0
                            average_discharge_quantity = 0
                        loadQuantity = average_load_quantity
                        dischargeQuantity = average_discharge_quantity

                expected_occupancy = (dischargeQuantity - loadQuantity + float(actual_occupancy)) / float(operational_capacity) * 100
                expected_teu = (dischargeQuantity - loadQuantity + float(actual_occupancy))
                updated_values[vessel_service] = (loadQuantity, dischargeQuantity)
                actual_occupancy = float(actual_occupancy) + dischargeQuantity - loadQuantity
                self.insert_or_update_calculated_data(etadate, vessel_service, vessel_name, loadQuantity, dischargeQuantity, expected_teu, expected_occupancy, PrimaryKey)

                self.display_forecast_data()
            
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error fetching data and calculating expected occupancy: {e}")
    
    def display_forecast_data(self):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                        SELECT etadate, vessel_service, vessel_name, PrimaryKey, load_quantity, discharge_quantity, expected_teu, expected_occupancy 
                        FROM Data_forecast""")
            forecast_data = cursor.fetchall()
            
            self.ui.tableWidget.clearContents()
            self.ui.tableWidget.setRowCount(len(forecast_data))
            
            for row_index, row_data in enumerate(forecast_data):
                etadate, vessel_service, vessel_name, PrimaryKey, load_quantity, discharge_quantity, expected_teu, expected_occupancy = row_data
                self.ui.tableWidget.setItem(row_index, 0, QTableWidgetItem(etadate))
                self.ui.tableWidget.setItem(row_index, 1, QTableWidgetItem(vessel_service))
                self.ui.tableWidget.setItem(row_index, 2, QTableWidgetItem(vessel_name))
                self.ui.tableWidget.setItem(row_index, 3, QTableWidgetItem(str(round(load_quantity, 2))))
                self.ui.tableWidget.setItem(row_index, 4, QTableWidgetItem(str(round(discharge_quantity, 2))))
                self.ui.tableWidget.setItem(row_index, 5, QTableWidgetItem(str(round(expected_teu, 2))))
                self.ui.tableWidget.setItem(row_index, 6, QTableWidgetItem(str(round(expected_occupancy, 2))))
            
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Error displaying forecast data: {e}")
    
    def recalculate_and_reload(self):
        try:
            conn = self.create_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT TOP 1 actual_occupancy, operational_capacity 
                FROM state 
                ORDER BY change_date DESC 
            """)
            actual_occupancy, operational_capacity = cursor.fetchone()

            self.ui.tableWidget_2.setItem(0, 0, QTableWidgetItem(str(actual_occupancy)))
            self.ui.tableWidget_2.setItem(0, 1, QTableWidgetItem(str(operational_capacity)))

            cursor.close()
            conn.close()

            self.fetch_data_and_calculate_expected_occupancy()

        except Exception as e:
            print(f"Error recalculating and reloading data: {e}")
    
    def handle_item_changed(self, item):
        row = item.row()
        col = item.column()
        if row == 0 and (col == 0 or col == 1):  
            try:
                actual_occupancy_item = self.ui.tableWidget_2.item(0, 0)
                operational_capacity_item = self.ui.tableWidget_2.item(0, 1)

                if actual_occupancy_item is not None and operational_capacity_item is not None:
                    actual_occupancy = float(actual_occupancy_item.text())
                    operational_capacity = float(operational_capacity_item.text())
                    self.ui.tableWidget_2.itemChanged.disconnect(self.handle_item_changed)

                    if not self.user_initiated_change:
                        self.user_initiated_change = True
                        self.update_db(actual_occupancy, operational_capacity)
                        self.user_initiated_change = False
                    self.recalculate_and_reload()
                    self.ui.tableWidget_2.itemChanged.connect(self.handle_item_changed)

            except ValueError:
                print("Invalid input. Please enter numeric values.")
            except Exception as e:
                print(f"Error handling item change: {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
